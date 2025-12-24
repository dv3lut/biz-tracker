from __future__ import annotations

from datetime import datetime
from types import SimpleNamespace
from uuid import uuid4

from openpyxl import load_workbook

from app.services import export_service


def _load_rows(buffer):
    workbook = load_workbook(buffer)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    return sheet, rows


def _make_establishment(**overrides):
    defaults = dict(
        date_creation=datetime(2024, 1, 1).date(),
        siret="12345678901234",
        name="Chez Test",
        numero_voie="10",
        indice_repetition=None,
        type_voie="Rue",
        libelle_voie="des Tests",
        code_postal="75000",
        libelle_commune="Paris",
        libelle_commune_etranger=None,
        code_pays="FR",
        categorie_entreprise="PME",
        categorie_juridique="5498",
        naf_code="5610A",
        naf_libelle="Restauration",
        google_place_id="place-123",
        google_place_url="https://maps.google.com/abc",
        google_match_confidence=0.9123,
        google_check_status="found",
        google_last_checked_at=datetime(2024, 1, 2, 10, 0, 0),
        google_last_found_at=datetime(2024, 1, 2, 11, 0, 0),
        google_listing_origin_at=datetime(2024, 1, 1, 9, 0, 0),
        google_listing_origin_source="google",
        google_listing_age_status="recent_creation",
        google_contact_phone=None,
        google_contact_email=None,
        google_contact_website=None,
        created_run_id=uuid4(),
        last_run_id=uuid4(),
        first_seen_at=datetime(2024, 1, 1, 8, 0, 0),
        last_seen_at=datetime(2024, 1, 3, 8, 0, 0),
    )
    defaults.update(overrides)
    return SimpleNamespace(**defaults)


def test_build_google_places_workbook_admin_mode_sets_hyperlinks():
    buffer = export_service.build_google_places_workbook([_make_establishment()])

    sheet, rows = _load_rows(buffer)

    assert sheet.title == "Google Places (admin)"
    assert len(rows) == 2
    siret_cell = sheet.cell(row=2, column=2)
    assert siret_cell.hyperlink and siret_cell.hyperlink.target.endswith("12345678901234")
    google_cell = sheet.cell(row=2, column=10)
    assert google_cell.hyperlink and "maps.google.com" in google_cell.hyperlink.target


def test_build_google_places_workbook_client_mode_uses_lookup():
    establishment = _make_establishment(google_listing_age_status="not_recent_creation")
    lookup = {"5610A": ("Restauration", "Bistrot")}  # category, subcategory

    buffer = export_service.build_google_places_workbook(
        [establishment],
        mode="client",
        subcategory_lookup=lookup,
        listing_statuses=["recent_creation", "not_recent_creation"],
    )
    sheet, rows = _load_rows(buffer)

    assert sheet.title == "Google Places (clients)"
    assert rows[0][:5] == (
        "Nom",
        "Adresse",
        "Catégorie",
        "Lien Google",
        "Statut fiche Google",
    )
    assert rows[1][0] == "Chez Test"
    assert rows[1][1] == "10 Rue des Tests, 75000 Paris"
    assert rows[1][2] == "Restauration"
    assert rows[1][4] == "Modification administrative récente"  # Label client, pas admin
    link_cell = sheet.cell(row=2, column=4)
    assert link_cell.hyperlink and "maps.google.com" in link_cell.hyperlink.target


def test_client_export_hides_status_when_single_filter():
    establishment = _make_establishment()

    buffer = export_service.build_google_places_workbook(
        [establishment],
        mode="client",
        subcategory_lookup=None,
        listing_statuses=["recent_creation"],
    )
    _, rows = _load_rows(buffer)

    expected_headers = ("Nom", "Adresse", "Catégorie", "Lien Google")
    assert rows[0][:4] == expected_headers
    assert rows[1][:4] == (
        establishment.name,
        "10 Rue des Tests, 75000 Paris",
        establishment.naf_libelle,
        establishment.google_place_url,
    )
    assert rows[1][4] is None


def test_build_alerts_workbook_includes_payload_and_links():
    establishment = _make_establishment()
    alert = SimpleNamespace(
        establishment=establishment,
        created_at=datetime(2024, 1, 5, 8, 0, 0),
        sent_at=None,
        run_id=uuid4(),
        recipients=["ops@example.com"],
        payload={"key": "value"},
        run=SimpleNamespace(scope_key="restaurants"),
    )

    buffer = export_service.build_alerts_workbook([alert])
    sheet, rows = _load_rows(buffer)

    assert sheet.title == "Alertes"
    assert rows[1][2] is None  # sent_at
    assert "ops@example.com" in rows[1][14]
    assert sheet.cell(row=2, column=4).hyperlink is not None


def test_export_helpers_normalize_values():
    establishment = _make_establishment(indice_repetition="BIS")

    assert export_service._format_datetime(None) is None
    assert export_service._format_date(None) is None
    assert export_service._compose_address(establishment) == "10 BIS Rue des Tests"
    assert export_service._compose_full_address(establishment) == "10 BIS Rue des Tests, 75000 Paris"
    assert export_service._format_date(datetime(2024, 1, 1)) == "2024-01-01T00:00:00"
    assert export_service._format_datetime(datetime(2024, 1, 1, 12, 0, 0)) == "2024-01-01T12:00:00"
    lookup = {"5610A": ("Restauration", "Bistrot")}
    assert export_service._resolve_category_columns("5610A", "Fallback", lookup) == ("Restauration", "Bistrot")
    assert export_service._resolve_category_columns(None, "Fallback", None) == ("Fallback", None)


def test_compose_full_address_handles_missing_segments():
    establishment = _make_establishment(
        numero_voie=None,
        type_voie=None,
        libelle_voie=None,
        code_postal=None,
        libelle_commune=None,
        libelle_commune_etranger=None,
    )

    assert export_service._compose_full_address(establishment) is None
