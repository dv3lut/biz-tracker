"""Utilities to export data snapshots (Excel, CSV, etc.)."""
from __future__ import annotations

import json
from io import BytesIO
from typing import Iterable, Literal, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.db import models
from app.utils.google_listing import describe_listing_age_status
from app.utils.urls import build_annuaire_etablissement_url


def _format_datetime(value: object | None) -> str | None:
    if value is None:
        return None
    return getattr(value, "isoformat", lambda: str(value))()


def _compose_address(establishment: models.Establishment) -> str:
    parts = [
        establishment.numero_voie,
        establishment.indice_repetition,
        establishment.type_voie,
        establishment.libelle_voie,
    ]
    return " ".join(filter(None, parts)).strip()


def _compose_full_address(establishment: models.Establishment) -> str | None:
    street = _compose_address(establishment)
    city_parts = [
        establishment.code_postal,
        establishment.libelle_commune or establishment.libelle_commune_etranger,
    ]
    city_line = " ".join(filter(None, city_parts)).strip()
    segments = [segment for segment in [street, city_line] if segment]
    if not segments:
        return None
    return ", ".join(segments)


def _format_date(value: object | None) -> str | None:
    if not value:
        return None
    formatter = getattr(value, "isoformat", None)
    if callable(formatter):
        return formatter()
    return str(value)


def _resolve_category_columns(
    naf_code: str | None,
    naf_label: str | None,
    lookup: Mapping[str, tuple[str | None, str | None]] | None,
) -> tuple[str | None, str | None]:
    if not naf_code or not lookup:
        return naf_label, None
    token = naf_code.strip().upper()
    if not token:
        return naf_label, None
    category_name, subcategory_name = lookup.get(token, (None, None))
    return category_name or naf_label, subcategory_name


def _apply_hyperlink(sheet, row_index: int, column_index: int, url: str | None) -> None:
    if not url:
        return
    cell = sheet.cell(row=row_index, column=column_index)
    cell.hyperlink = url
    cell.style = "Hyperlink"


def build_google_places_workbook(
    establishments: Iterable[models.Establishment],
    *,
    mode: Literal["admin", "client"] = "admin",
    subcategory_lookup: Mapping[str, tuple[str | None, str | None]] | None = None,
    listing_statuses: Sequence[str] | None = None,
) -> BytesIO:
    """Generate an Excel workbook listing establishments enriched with Google Places."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Google Places (clients)" if mode == "client" else "Google Places (admin)"
    if mode == "client":
        selected_statuses = list(listing_statuses or [])
        include_listing_status = len(selected_statuses) != 1
        headers = [
            "Nom",
            "Adresse",
            "Catégorie",
            "Sous-catégorie",
            "Lien Google",
        ]
        if include_listing_status:
            headers.append("Statut fiche Google")
    else:
        headers = [
            "Date création",
            "SIRET",
            "Nom",
            "Adresse",
            "Code postal",
            "Commune",
            "Pays",
            "Google Place ID",
            "Google Place URL",
            "Score correspondance",
            "Statut Google",
            "Dernière vérification",
            "Dernière détection",
            "Origine fiche Google",
            "Statut fiche Google",
            "Run de création",
            "Run le plus récent",
            "Vu en premier",
            "Vu en dernier",
        ]
    sheet.append(headers)

    if mode == "client":
        link_column_index = headers.index("Lien Google") + 1
    else:
        link_column_index = None

    for establishment in establishments:
        creation_date = _format_date(establishment.date_creation)
        address = _compose_address(establishment)
        commune = establishment.libelle_commune or establishment.libelle_commune_etranger
        google_url = establishment.google_place_url
        if mode == "client":
            full_address = _compose_full_address(establishment)
            category_name, subcategory_name = _resolve_category_columns(
                establishment.naf_code,
                establishment.naf_libelle,
                subcategory_lookup,
            )
            row = [
                establishment.name,
                full_address,
                category_name,
                subcategory_name,
                google_url,
            ]
            if include_listing_status:
                row.append(describe_listing_age_status(establishment.google_listing_age_status))
            sheet.append(row)
            _apply_hyperlink(sheet, sheet.max_row, link_column_index, google_url)
            continue

        sheet.append(
            [
                creation_date,
                establishment.siret,
                establishment.name,
                address,
                establishment.code_postal,
                commune,
                establishment.code_pays,
                establishment.google_place_id,
                google_url,
                round(establishment.google_match_confidence, 3) if establishment.google_match_confidence is not None else None,
                establishment.google_check_status,
                _format_datetime(establishment.google_last_checked_at),
                _format_datetime(establishment.google_last_found_at),
                _format_datetime(establishment.google_listing_origin_at),
                describe_listing_age_status(establishment.google_listing_age_status),
                str(establishment.created_run_id) if establishment.created_run_id else None,
                str(establishment.last_run_id) if establishment.last_run_id else None,
                _format_datetime(establishment.first_seen_at),
                _format_datetime(establishment.last_seen_at),
            ]
        )
    row_idx = sheet.max_row
    _apply_hyperlink(sheet, row_idx, 2, build_annuaire_etablissement_url(establishment.siret))
    _apply_hyperlink(sheet, row_idx, 10, google_url)

    sheet.freeze_panes = "A2"

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                value_length = len(str(cell.value)) if cell.value is not None else 0
            except Exception:  # pragma: no cover - defensive
                value_length = 0
            max_length = max(max_length, value_length)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_alerts_workbook(alerts: Iterable[models.Alert]) -> BytesIO:
    """Generate an Excel workbook listing alerts filtered by business creation date."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Alertes"
    headers = [
        "Date création",
        "Date alerte",
        "Date envoi",
        "SIRET",
        "Nom",
        "Adresse",
        "Code postal",
        "Commune",
        "Pays",
        "NAF",
        "Catégorie entreprise",
        "Catégorie juridique",
        "Run ID",
        "Scope",
        "Destinataires",
        "Payload",
    ]
    sheet.append(headers)

    for alert in alerts:
        establishment = alert.establishment
        if establishment is None:
            continue

        payload_str = json.dumps(alert.payload or {}, ensure_ascii=False)
        recipients = ", ".join(alert.recipients or [])
        sheet.append(
            [
                establishment.date_creation.isoformat() if establishment.date_creation else None,
                _format_datetime(alert.created_at),
                _format_datetime(alert.sent_at),
                establishment.siret,
                establishment.name,
                _compose_address(establishment),
                establishment.code_postal,
                establishment.libelle_commune or establishment.libelle_commune_etranger,
                establishment.code_pays,
                establishment.naf_code,
                establishment.categorie_entreprise,
                establishment.categorie_juridique,
                str(alert.run_id) if alert.run_id else None,
                alert.run.scope_key if alert.run else None,
                recipients,
                payload_str,
            ]
        )
        _apply_hyperlink(sheet, sheet.max_row, 4, build_annuaire_etablissement_url(establishment.siret))

    sheet.freeze_panes = "A2"

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                value_length = len(str(cell.value)) if cell.value is not None else 0
            except Exception:  # pragma: no cover - defensive
                value_length = 0
            max_length = max(max_length, value_length)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
