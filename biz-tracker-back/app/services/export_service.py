"""Utilities to export data snapshots (Excel, CSV, etc.)."""
from __future__ import annotations

import csv
import json
from io import BytesIO
from io import StringIO
from typing import Iterable, Literal, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.db import models
from app.services.alerts.email_renderer import get_client_listing_status_label
from app.utils.google_listing import describe_listing_age_status
from app.utils.urls import build_annuaire_etablissement_url


def _format_datetime(value: object | None) -> str | None:
    if value is None:
        return None
    return getattr(value, "isoformat", lambda: str(value))()


def _compose_address(establishment: models.Establishment) -> str:
    parts = [
        establishment.numero_voie,
        establishment.indice_repetition,
        establishment.type_voie,
        establishment.libelle_voie,
    ]
    return " ".join(filter(None, parts)).strip()


def _compose_full_address(establishment: models.Establishment) -> str | None:
    street = _compose_address(establishment)
    city_parts = [
        establishment.code_postal,
        establishment.libelle_commune or establishment.libelle_commune_etranger,
    ]
    city_line = " ".join(filter(None, city_parts)).strip()
    segments = [segment for segment in [street, city_line] if segment]
    if not segments:
        return None
    return ", ".join(segments)


def _format_date(value: object | None) -> str | None:
    if not value:
        return None
    formatter = getattr(value, "isoformat", None)
    if callable(formatter):
        return formatter()
    return str(value)


def _resolve_category_columns(
    naf_code: str | None,
    naf_label: str | None,
    lookup: Mapping[str, tuple[str | None, str | None]] | None,
) -> tuple[str | None, str | None]:
    if not naf_code or not lookup:
        return naf_label, None
    token = naf_code.strip().upper()
    if not token:
        return naf_label, None
    category_name, subcategory_name = lookup.get(token, (None, None))
    return category_name or naf_label, subcategory_name


def _apply_hyperlink(sheet, row_index: int, column_index: int, url: str | None) -> None:
    if not url:
        return
    cell = sheet.cell(row=row_index, column=column_index)
    cell.hyperlink = url
    cell.style = "Hyperlink"


def build_google_places_workbook(
    establishments: Iterable[models.Establishment],
    *,
    mode: Literal["admin", "client"] = "admin",
    subcategory_lookup: Mapping[str, tuple[str | None, str | None]] | None = None,
    listing_statuses: Sequence[str] | None = None,
) -> BytesIO:
    """Generate an Excel workbook listing establishments enriched with Google Places."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Google Places (clients)" if mode == "client" else "Google Places (admin)"
    if mode == "client":
        selected_statuses = list(listing_statuses or [])
        include_listing_status = len(selected_statuses) != 1
        headers = [
            "Nom",
            "Adresse",
            "Catégorie",
            "Lien Google",
        ]
        if include_listing_status:
            headers.append("Statut fiche Google")
    else:
        headers = [
            "Date création",
            "SIRET",
            "Nom",
            "Adresse",
            "Code postal",
            "Commune",
            "Pays",
            "Google Place ID",
            "Google Place URL",
            "Score correspondance",
            "Statut Google",
            "Dernière vérification",
            "Dernière détection",
            "Origine fiche Google",
            "Statut fiche Google",
            "Run de création",
            "Run le plus récent",
            "Vu en premier",
            "Vu en dernier",
        ]
    sheet.append(headers)

    if mode == "client":
        link_column_index = headers.index("Lien Google") + 1
    else:
        link_column_index = None

    for establishment in establishments:
        creation_date = _format_date(establishment.date_creation)
        address = _compose_address(establishment)
        commune = establishment.libelle_commune or establishment.libelle_commune_etranger
        google_url = establishment.google_place_url
        if mode == "client":
            full_address = _compose_full_address(establishment)
            category_name, _subcategory_name = _resolve_category_columns(
                establishment.naf_code,
                establishment.naf_libelle,
                subcategory_lookup,
            )
            row = [
                establishment.name,
                full_address,
                category_name,
                google_url,
            ]
            if include_listing_status:
                # Utiliser le label client pour les exports clients
                row.append(get_client_listing_status_label(establishment.google_listing_age_status))
            sheet.append(row)
            _apply_hyperlink(sheet, sheet.max_row, link_column_index, google_url)
            continue

        sheet.append(
            [
                creation_date,
                establishment.siret,
                establishment.name,
                address,
                establishment.code_postal,
                commune,
                establishment.code_pays,
                establishment.google_place_id,
                google_url,
                round(establishment.google_match_confidence, 3) if establishment.google_match_confidence is not None else None,
                establishment.google_check_status,
                _format_datetime(establishment.google_last_checked_at),
                _format_datetime(establishment.google_last_found_at),
                _format_datetime(establishment.google_listing_origin_at),
                describe_listing_age_status(establishment.google_listing_age_status),
                str(establishment.created_run_id) if establishment.created_run_id else None,
                str(establishment.last_run_id) if establishment.last_run_id else None,
                _format_datetime(establishment.first_seen_at),
                _format_datetime(establishment.last_seen_at),
            ]
        )
    row_idx = sheet.max_row
    _apply_hyperlink(sheet, row_idx, 2, build_annuaire_etablissement_url(establishment.siret))
    _apply_hyperlink(sheet, row_idx, 10, google_url)

    sheet.freeze_panes = "A2"

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                value_length = len(str(cell.value)) if cell.value is not None else 0
            except Exception:  # pragma: no cover - defensive
                value_length = 0
            max_length = max(max_length, value_length)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_alerts_workbook(alerts: Iterable[models.Alert]) -> BytesIO:
    """Generate an Excel workbook listing alerts filtered by business creation date."""

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Alertes"
    headers = [
        "Date création",
        "Date alerte",
        "Date envoi",
        "SIRET",
        "Nom",
        "Adresse",
        "Code postal",
        "Commune",
        "Pays",
        "NAF",
        "Catégorie entreprise",
        "Catégorie juridique",
        "Run ID",
        "Scope",
        "Destinataires",
        "Payload",
    ]
    sheet.append(headers)

    for alert in alerts:
        establishment = alert.establishment
        if establishment is None:
            continue

        payload_str = json.dumps(alert.payload or {}, ensure_ascii=False)
        recipients = ", ".join(alert.recipients or [])
        sheet.append(
            [
                establishment.date_creation.isoformat() if establishment.date_creation else None,
                _format_datetime(alert.created_at),
                _format_datetime(alert.sent_at),
                establishment.siret,
                establishment.name,
                _compose_address(establishment),
                establishment.code_postal,
                establishment.libelle_commune or establishment.libelle_commune_etranger,
                establishment.code_pays,
                establishment.naf_code,
                establishment.categorie_entreprise,
                establishment.categorie_juridique,
                str(alert.run_id) if alert.run_id else None,
                alert.run.scope_key if alert.run else None,
                recipients,
                payload_str,
            ]
        )
        _apply_hyperlink(sheet, sheet.max_row, 4, build_annuaire_etablissement_url(establishment.siret))

    sheet.freeze_panes = "A2"

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                value_length = len(str(cell.value)) if cell.value is not None else 0
            except Exception:  # pragma: no cover - defensive
                value_length = 0
            max_length = max(max_length, value_length)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def build_alerts_csv(
    alerts: Iterable[models.Alert],
    *,
    delimiter: str = ";",
    establishments_by_siret: Mapping[str, object] | None = None,
    scope_key: str | None = None,
) -> bytes:
    """Generate a CSV export for alerts.

    Mirrors the columns of :func:`build_alerts_workbook` so that emails and UI exports
    keep a consistent schema.
    """

    headers = [
        "Date création",
        "Date alerte",
        "Date envoi",
        "SIRET",
        "Nom",
        "Adresse",
        "Code postal",
        "Commune",
        "Pays",
        "NAF",
        "Catégorie entreprise",
        "Catégorie juridique",
        "Run ID",
        "Scope",
        "Destinataires",
        "Payload",
    ]

    def normalize(value: object | None) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value
        return str(value)

    buffer = StringIO(newline="")
    writer = csv.writer(buffer, delimiter=delimiter)
    writer.writerow(headers)

    for alert in alerts:
        establishment = getattr(alert, "establishment", None)
        if establishment is None and establishments_by_siret is not None:
            establishment = establishments_by_siret.get(getattr(alert, "siret", ""))
        if establishment is None:
            continue

        payload_str = json.dumps(getattr(alert, "payload", None) or {}, ensure_ascii=False)
        recipients = ", ".join(getattr(alert, "recipients", None) or [])

        writer.writerow(
            [
                normalize(_format_date(getattr(establishment, "date_creation", None))),
                normalize(_format_datetime(getattr(alert, "created_at", None))),
                normalize(_format_datetime(getattr(alert, "sent_at", None))),
                normalize(getattr(establishment, "siret", None)),
                normalize(getattr(establishment, "name", None)),
                normalize(_compose_address(establishment)),
                normalize(getattr(establishment, "code_postal", None)),
                normalize(getattr(establishment, "libelle_commune", None) or getattr(establishment, "libelle_commune_etranger", None)),
                normalize(getattr(establishment, "code_pays", None)),
                normalize(getattr(establishment, "naf_code", None)),
                normalize(getattr(establishment, "categorie_entreprise", None)),
                normalize(getattr(establishment, "categorie_juridique", None)),
                normalize(str(getattr(alert, "run_id", "")) if getattr(alert, "run_id", None) else ""),
                normalize(scope_key or getattr(getattr(alert, "run", None), "scope_key", None)),
                normalize(recipients),
                normalize(payload_str),
            ]
        )

    # utf-8-sig: include BOM so Excel on Windows detects UTF-8 reliably.
    return buffer.getvalue().encode("utf-8-sig")


MONTH_LABELS_FR_EXPORT: dict[int, str] = {
    1: "janvier", 2: "février", 3: "mars", 4: "avril", 5: "mai", 6: "juin",
    7: "juillet", 8: "août", 9: "septembre", 10: "octobre", 11: "novembre", 12: "décembre",
}


def _format_month_year(value: object | None) -> str | None:
    """Return 'Mois YYYY' string for a date-like value."""
    if not value:
        return None
    month = getattr(value, "month", None)
    year = getattr(value, "year", None)
    if month and year:
        return f"{MONTH_LABELS_FR_EXPORT[month].capitalize()} {year}"
    return None


def build_alerts_client_csv(
    alerts: Iterable[models.Alert],
    *,
    delimiter: str = ";",
    establishments_by_siret: Mapping[str, object] | None = None,
) -> bytes:
    """Generate a client-safe CSV export for alerts.

    Only includes non-sensitive fields suitable for customers.
    """

    headers = [
        "Mois/Année création",
        "Date alerte",
        "Nom",
        "Adresse complète",
        "Code postal",
        "Commune",
        "Pays",
        "Catégorie",
        "Statut fiche Google",
        "Fiche Google",
        "Entreprise individuelle",
        "Dirigeant(s)",
    ]

    def normalize(value: object | None) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value
        return str(value)

    buffer = StringIO(newline="")
    writer = csv.writer(buffer, delimiter=delimiter)
    writer.writerow(headers)

    for alert in alerts:
        establishment = getattr(alert, "establishment", None)
        if establishment is None and establishments_by_siret is not None:
            establishment = establishments_by_siret.get(getattr(alert, "siret", ""))
        if establishment is None:
            continue

        # Entreprise individuelle
        sole = getattr(establishment, "is_sole_proprietorship", None)
        if sole is True:
            sole_str = "Oui"
        elif sole is False:
            sole_str = "Non"
        else:
            # Fallback via categorie_juridique si la propriété n'est pas disponible
            from app.utils.business_types import is_individual_company  # local to avoid circular
            cj = getattr(establishment, "categorie_juridique", None)
            sole_str = "Oui" if is_individual_company(cj) else "Non"

        # Dirigeants (personnes physiques uniquement)
        directors = getattr(establishment, "directors", None) or []
        physical_directors = [d for d in directors if getattr(d, "is_physical_person", False)]
        directors_parts: list[str] = []
        for d in physical_directors:
            first_names = (getattr(d, "first_names", None) or "").strip()
            last_name = (getattr(d, "last_name", None) or "").strip()
            quality = (getattr(d, "quality", None) or "").strip()
            if first_names and last_name:
                name_str = f"{first_names} {last_name.upper()}"
            elif last_name:
                name_str = last_name.upper()
            elif first_names:
                name_str = first_names
            else:
                name_str = quality
            directors_parts.append(f"{name_str} ({quality})" if quality and quality != name_str else name_str)
        directors_str = " | ".join(directors_parts)

        # Statut fiche Google (label client)
        listing_status = get_client_listing_status_label(
            getattr(establishment, "google_listing_age_status", None)
        )

        writer.writerow(
            [
                normalize(_format_month_year(getattr(establishment, "date_creation", None))),
                normalize(_format_datetime(getattr(alert, "created_at", None))),
                normalize(getattr(establishment, "name", None)),
                normalize(_compose_full_address(establishment)),
                normalize(getattr(establishment, "code_postal", None)),
                normalize(getattr(establishment, "libelle_commune", None) or getattr(establishment, "libelle_commune_etranger", None)),
                normalize(getattr(establishment, "code_pays", None)),
                normalize(getattr(establishment, "naf_libelle", None)),
                normalize(listing_status),
                normalize(getattr(establishment, "google_place_url", None)),
                sole_str,
                normalize(directors_str),
            ]
        )

    return buffer.getvalue().encode("utf-8-sig")
