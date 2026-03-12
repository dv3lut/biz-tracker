from __future__ import annotations

from datetime import datetime
from types import SimpleNamespace
from uuid import uuid4

from openpyxl import load_workbook

from app.services import export_service


def _load_rows(buffer):
    workbook = load_workbook(buffer)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    return sheet, rows


def _make_establishment(**overrides):
    defaults = dict(
        date_creation=datetime(2024, 1, 1).date(),
        siret="12345678901234",
        name="Chez Test",
        numero_voie="10",
        indice_repetition=None,
        type_voie="Rue",
        libelle_voie="des Tests",
        code_postal="75000",
        libelle_commune="Paris",
        libelle_commune_etranger=None,
        code_pays="FR",
        categorie_entreprise="PME",
        categorie_juridique="5498",
        naf_code="5610A",
        naf_libelle="Restauration",
        google_place_id="place-123",
        google_place_url="https://maps.google.com/abc",
        google_match_confidence=0.9123,
        google_check_status="found",
        google_last_checked_at=datetime(2024, 1, 2, 10, 0, 0),
        google_last_found_at=datetime(2024, 1, 2, 11, 0, 0),
        google_listing_origin_at=datetime(2024, 1, 1, 9, 0, 0),
        google_listing_origin_source="google",
        google_listing_age_status="recent_creation",
        google_contact_phone="+33102030405",
        google_contact_email="contact@chez-test.fr",
        google_contact_website="https://chez-test.fr",
        website_scraped_at=datetime(2024, 1, 2, 12, 0, 0),
        website_scraped_mobile_phones="+33601020304|+33605060708",
        website_scraped_national_phones="+33111223344|+33955667788",
        website_scraped_international_phones="+442079460958",
        website_scraped_emails="hello@chez-test.fr|booking@chez-test.fr",
        website_scraped_facebook="https://facebook.com/cheztest",
        website_scraped_instagram="https://instagram.com/cheztest",
        website_scraped_twitter="https://x.com/cheztest",
        website_scraped_linkedin="https://linkedin.com/company/cheztest",
        created_run_id=uuid4(),
        last_run_id=uuid4(),
        first_seen_at=datetime(2024, 1, 1, 8, 0, 0),
        last_seen_at=datetime(2024, 1, 3, 8, 0, 0),
    )
    defaults.update(overrides)
    return SimpleNamespace(**defaults)


def test_build_google_places_workbook_admin_mode_sets_hyperlinks():
    buffer = export_service.build_google_places_workbook([_make_establishment()])

    sheet, rows = _load_rows(buffer)

    assert sheet.title == "Google Places (admin)"
    assert len(rows) == 2
    siret_col = rows[0].index("SIRET") + 1
    google_col = rows[0].index("Google Place URL") + 1
    siret_cell = sheet.cell(row=2, column=siret_col)
    assert siret_cell.hyperlink and siret_cell.hyperlink.target.endswith("12345678901234")
    google_cell = sheet.cell(row=2, column=google_col)
    assert google_cell.hyperlink and "maps.google.com" in google_cell.hyperlink.target


def test_build_google_places_workbook_admin_mode_includes_google_scraping_and_naf_columns():
    establishment = _make_establishment(naf_code="5610A", naf_libelle="Restauration")
    lookup = {"5610A": ("Restauration commerciale", "Restauration traditionnelle")}

    buffer = export_service.build_google_places_workbook(
        [establishment],
        mode="admin",
        subcategory_lookup=lookup,
    )
    _, rows = _load_rows(buffer)

    headers = rows[0]
    values = rows[1]
    assert "Téléphone Google" in headers
    assert "Email Google" in headers
    assert "Site web Google" in headers
    assert "Téléphones mobiles (site)" in headers
    assert "Téléphones fixes (site)" in headers
    assert "Téléphones internationaux (site)" in headers
    assert "Emails (site)" in headers
    assert "Catégorie NAF" in headers
    assert "Sous-catégorie NAF" in headers
    value_by_header = dict(zip(headers, values))
    assert value_by_header["Téléphone Google"] == "+33102030405"
    assert value_by_header["Email Google"] == "contact@chez-test.fr"
    assert value_by_header["Site web Google"] == "https://chez-test.fr"
    assert value_by_header["Téléphones mobiles (site)"] == "+33601020304, +33605060708"
    assert value_by_header["Téléphones fixes (site)"] == "+33111223344, +33955667788"
    assert value_by_header["Téléphones internationaux (site)"] == "+442079460958"
    assert value_by_header["Emails (site)"] == "hello@chez-test.fr, booking@chez-test.fr"
    assert value_by_header["Facebook (site)"] == "https://facebook.com/cheztest"
    assert value_by_header["Instagram (site)"] == "https://instagram.com/cheztest"
    assert value_by_header["Twitter/X (site)"] == "https://x.com/cheztest"
    assert value_by_header["LinkedIn (site)"] == "https://linkedin.com/company/cheztest"
    assert value_by_header["Catégorie NAF"] == "Restauration commerciale"
    assert value_by_header["Sous-catégorie NAF"] == "Restauration traditionnelle"


def test_build_google_places_workbook_client_mode_uses_lookup():
    establishment = _make_establishment(google_listing_age_status="not_recent_creation")
    lookup = {"5610A": ("Restauration", "Bistrot")}  # category, subcategory

    buffer = export_service.build_google_places_workbook(
        [establishment],
        mode="client",
        subcategory_lookup=lookup,
        listing_statuses=["recent_creation", "not_recent_creation"],
    )
    sheet, rows = _load_rows(buffer)

    assert sheet.title == "Google Places (clients)"
    assert rows[0][:5] == (
        "Nom",
        "Adresse",
        "Catégorie",
        "Lien Google",
        "Statut fiche Google",
    )
    assert rows[1][0] == "Chez Test"
    assert rows[1][1] == "10 Rue des Tests, 75000 Paris"
    assert rows[1][2] == "Restauration"
    assert rows[1][4] == "Modification administrative récente"  # Label client, pas admin
    link_cell = sheet.cell(row=2, column=4)
    assert link_cell.hyperlink and "maps.google.com" in link_cell.hyperlink.target


def test_client_export_hides_status_when_single_filter():
    establishment = _make_establishment()

    buffer = export_service.build_google_places_workbook(
        [establishment],
        mode="client",
        subcategory_lookup=None,
        listing_statuses=["recent_creation"],
    )
    _, rows = _load_rows(buffer)

    expected_headers = ("Nom", "Adresse", "Catégorie", "Lien Google")
    assert rows[0][:4] == expected_headers
    assert rows[1][:4] == (
        establishment.name,
        "10 Rue des Tests, 75000 Paris",
        establishment.naf_libelle,
        establishment.google_place_url,
    )
    assert len(rows[1]) == 4


def test_build_alerts_workbook_includes_payload_and_links():
    establishment = _make_establishment()
    alert = SimpleNamespace(
        establishment=establishment,
        created_at=datetime(2024, 1, 5, 8, 0, 0),
        sent_at=None,
        run_id=uuid4(),
        recipients=["ops@example.com"],
        payload={"key": "value"},
        run=SimpleNamespace(scope_key="restaurants"),
    )

    buffer = export_service.build_alerts_workbook([alert])
    sheet, rows = _load_rows(buffer)

    assert sheet.title == "Alertes"
    assert rows[1][2] is None  # sent_at
    assert "ops@example.com" in rows[1][14]
    assert sheet.cell(row=2, column=4).hyperlink is not None


def test_build_alerts_csv_includes_headers_and_payload():
    establishment = _make_establishment()
    alert = SimpleNamespace(
        establishment=establishment,
        created_at=datetime(2024, 1, 5, 8, 0, 0),
        sent_at=None,
        run_id=uuid4(),
        recipients=["ops@example.com"],
        payload={"key": "value"},
        run=SimpleNamespace(scope_key="restaurants"),
    )

    content = export_service.build_alerts_csv([alert]).decode("utf-8-sig")
    lines = [line for line in content.splitlines() if line]

    assert lines[0].startswith("Date création;")
    assert "ops@example.com" in lines[1]
    assert '"key": "value"' in lines[1] or "key" in lines[1]


def test_build_alerts_client_csv_is_minimal_and_safe():
    from types import SimpleNamespace as SN

    region = SN(name="Île-de-France")
    director = SN(
        is_physical_person=True,
        first_names="Jean",
        last_name="Dupont",
        quality="Gérant",
    )
    establishment = _make_establishment()
    establishment.is_sole_proprietorship = False
    establishment.directors = [director]
    alert = SimpleNamespace(
        establishment=establishment,
        created_at=datetime(2024, 1, 5, 8, 0, 0),
        sent_at=None,
        run_id=uuid4(),
        recipients=["ops@example.com"],
        payload={"secret": "value"},
        run=SimpleNamespace(scope_key="restaurants"),
    )

    content = export_service.build_alerts_client_csv([alert]).decode("utf-8-sig")
    header, row = [line for line in content.splitlines() if line][:2]

    expected_header = (
        "Mois/Année création;Date alerte;Nom;Adresse complète;"
        "Code postal;Commune;Pays;Catégorie;Statut fiche Google;"
        "Fiche Google;Entreprise individuelle;Dirigeant(s)"
    )
    assert header == expected_header
    assert "ops@example.com" not in row
    assert "secret" not in row
    # Mois/Année et non date précise
    assert "Janvier 2024" in row
    assert "2024-01-01" not in row
    # Dirigeant format
    assert "Jean DUPONT" in row
    assert "Gérant" in row
    # Entreprise individuelle
    assert "Non" in row
    # Statut Google
    assert "récente" in row.lower() or "création" in row.lower()


def test_export_helpers_normalize_values():
    establishment = _make_establishment(indice_repetition="BIS")

    assert export_service._format_datetime(None) is None
    assert export_service._format_date(None) is None
    assert export_service._compose_address(establishment) == "10 BIS Rue des Tests"
    assert export_service._compose_full_address(establishment) == "10 BIS Rue des Tests, 75000 Paris"
    assert export_service._format_date(datetime(2024, 1, 1)) == "2024-01-01T00:00:00"
    assert export_service._format_datetime(datetime(2024, 1, 1, 12, 0, 0)) == "2024-01-01T12:00:00"
    lookup = {"5610A": ("Restauration", "Bistrot")}
    assert export_service._resolve_category_columns("5610A", "Fallback", lookup) == ("Restauration", "Bistrot")
    assert export_service._resolve_category_columns(None, "Fallback", None) == ("Fallback", None)


def test_compose_full_address_handles_missing_segments():
    establishment = _make_establishment(
        numero_voie=None,
        type_voie=None,
        libelle_voie=None,
        code_postal=None,
        libelle_commune=None,
        libelle_commune_etranger=None,
    )

    assert export_service._compose_full_address(establishment) is None


# ---------------------------------------------------------------------------
# Branches supplémentaires pour monter la couverture
# ---------------------------------------------------------------------------


def test_format_date_fallback_to_str():
    """_format_date doit retourner str(value) quand la valeur n'a pas d'isoformat."""
    assert export_service._format_date(20240101) == "20240101"


def test_resolve_category_columns_whitespace_naf_code():
    """_resolve_category_columns renvoie (naf_label, None) quand le code NAF est vide après strip."""
    result = export_service._resolve_category_columns("   ", "Fallback", {"X": ("Cat", "Sub")})
    assert result == ("Fallback", None)


def test_apply_hyperlink_skips_none_url():
    """build_google_places_workbook avec google_place_url=None ne doit pas lever d'exception."""
    establishment = _make_establishment(google_place_url=None, google_place_id=None)
    buffer = export_service.build_google_places_workbook([establishment])
    sheet, rows = _load_rows(buffer)
    # La ligne doit exister mais sans hyperlien sur la colonne URL
    google_col = rows[0].index("Google Place URL") + 1
    assert rows[1][google_col - 1] is None
    assert sheet.cell(row=2, column=google_col).hyperlink is None


def test_build_alerts_workbook_skips_alert_without_establishment():
    """Un alerte avec establishment=None est ignorée silencieusement dans le workbook."""
    alert_no_estab = SimpleNamespace(
        establishment=None,
        created_at=datetime(2024, 1, 5, 8, 0, 0),
        sent_at=None,
        run_id=uuid4(),
        recipients=[],
        payload={},
        run=SimpleNamespace(scope_key="test"),
    )
    buffer = export_service.build_alerts_workbook([alert_no_estab])
    _, rows = _load_rows(buffer)
    assert len(rows) == 1  # en-tête uniquement


def test_build_alerts_csv_establishments_by_siret_fallback():
    """build_alerts_csv utilise establishments_by_siret quand alert.establishment est None."""
    establishment = _make_establishment()
    alert = SimpleNamespace(
        establishment=None,
        siret="12345678901234",
        created_at=datetime(2024, 1, 5, 8, 0, 0),
        sent_at=None,
        run_id=uuid4(),
        recipients=["ops@example.com"],
        payload={},
        run=SimpleNamespace(scope_key="test"),
    )
    content = export_service.build_alerts_csv(
        [alert],
        establishments_by_siret={"12345678901234": establishment},
    ).decode("utf-8-sig")
    lines = [line for line in content.splitlines() if line]
    assert len(lines) == 2
    assert "Chez Test" in lines[1]


def test_build_alerts_csv_skips_alert_without_establishment():
    """build_alerts_csv ignore un alerte sans établissement ni fallback map."""
    alert = SimpleNamespace(
        establishment=None,
        siret="99999999999999",
        created_at=datetime(2024, 1, 5, 8, 0, 0),
        sent_at=None,
        run_id=uuid4(),
        recipients=[],
        payload={},
        run=None,
    )
    content = export_service.build_alerts_csv([alert]).decode("utf-8-sig")
    lines = [line for line in content.splitlines() if line]
    assert len(lines) == 1  # en-tête uniquement


def test_format_month_year_returns_none_without_month_attr():
    """_format_month_year renvoie None si la valeur n'a pas d'attribut month/year."""
    assert export_service._format_month_year("not-a-date") is None
    assert export_service._format_month_year(None) is None


def test_build_alerts_client_csv_skips_alert_without_establishment():
    """build_alerts_client_csv ignore un alerte sans établissement."""
    alert = SimpleNamespace(
        establishment=None,
        siret="99999999999999",
        created_at=datetime(2024, 1, 5, 8, 0, 0),
        sent_at=None,
        run_id=uuid4(),
        recipients=[],
        payload={},
        run=None,
    )
    content = export_service.build_alerts_client_csv([alert]).decode("utf-8-sig")
    lines = [line for line in content.splitlines() if line]
    assert len(lines) == 1  # en-tête uniquement


def test_build_alerts_client_csv_sole_proprietorship_true():
    """build_alerts_client_csv affiche 'Oui' pour une entreprise individuelle."""
    from types import SimpleNamespace as SN

    establishment = _make_establishment()
    establishment.is_sole_proprietorship = True
    establishment.directors = []
    alert = SimpleNamespace(
        establishment=establishment,
        created_at=datetime(2024, 1, 5, 8, 0, 0),
        sent_at=None,
        run_id=uuid4(),
        recipients=[],
        payload={},
        run=None,
    )
    content = export_service.build_alerts_client_csv([alert]).decode("utf-8-sig")
    _header, row = [line for line in content.splitlines() if line][:2]
    assert "Oui" in row


def test_build_alerts_client_csv_director_name_variants():
    """Formatage du nom du dirigeant : nom seul, prénom seul, ni l'un ni l'autre."""
    from types import SimpleNamespace as SN

    establishment = _make_establishment()
    establishment.is_sole_proprietorship = False
    establishment.directors = [
        SN(is_physical_person=True, first_names="", last_name="MARTIN", quality="Gérant"),
        SN(is_physical_person=True, first_names="Sophie", last_name="", quality="DG"),
        SN(is_physical_person=True, first_names="", last_name="", quality="Associé"),
        SN(is_physical_person=False, first_names="Corp SA", last_name="", quality=""),
    ]
    alert = SimpleNamespace(
        establishment=establishment,
        created_at=datetime(2024, 1, 5, 8, 0, 0),
        sent_at=None,
        run_id=uuid4(),
        recipients=[],
        payload={},
        run=None,
    )
    content = export_service.build_alerts_client_csv([alert]).decode("utf-8-sig")
    _header, row = [line for line in content.splitlines() if line][:2]
    assert "MARTIN" in row
    assert "Sophie" in row
    assert "Associé" in row
    assert "Corp SA" not in row  # personne morale exclue
